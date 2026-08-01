"""
Veo 3 Storyboard Station
excel_manager.py
"""
import openpyxl
import config
from pathlib import Path

# Các biến toàn cục quản lý trạng thái Excel hiện tại
_workbook = None
_sheet = None
_workbook_path = None
_current_loaded_type = None  # Biến theo dõi loại tab đang được nạp thực tế


def get_type_scenario_path(scenario_type):
    """
    Hàm bổ trợ: Trả về chính xác ĐƯỜNG DẪN FILE (Path) và khởi tạo file nếu chưa có.
    """
    if scenario_type == config.SCENARIO_IMAGE:
        file_path = config.IMAGE_SCENARIO_FILE
    elif scenario_type == config.SCENARIO_VIDEO:
        file_path = config.VIDEO_SCENARIO_FILE
    else:
        raise ValueError(f"Loại scenario không hợp lệ: {scenario_type}")

    # Nếu file chưa tồn tại, tạo mới hoàn toàn với cấu trúc chuẩn
    if not file_path.exists():
        wb = openpyxl.Workbook()

        # Tạo sheet chính "Script"
        ws_script = wb.active
        ws_script.title = config.SCENARIO_SHEET
        headers = sorted(config.EXCEL_COLUMNS.keys(), key=lambda k: config.EXCEL_COLUMNS[k])
        ws_script.append(headers)

        # Tạo sẵn sheet phụ "Templates" để tránh lỗi tra cứu sau này
        wb.create_sheet(title="Templates")

        wb.save(file_path)

    return file_path


def get_type_scenario(scenario_type):
    """
    Hàm phục vụ cho UI: Mở workbook và trả về đúng Worksheet kịch bản.
    Đồng thời đảm bảo đồng bộ hóa vào biến toàn cục.
    """
    return open_workbook(scenario_type)


# Workbook Management
# ==========================================================

def get_workbook():
    return _workbook


def open_workbook(scenario_type):
    """
    Mở file Excel tương ứng với Tab.
    Nếu chuyển đổi qua lại giữa Image và Video, hệ thống sẽ tự động đóng file cũ, mở file mới.
    """
    global _workbook, _sheet, _workbook_path, _current_loaded_type

    # CHỈ sử dụng lại cache nếu đúng loại scenario_type đang mở hiện tại
    if _workbook and _current_loaded_type == scenario_type:
        return _sheet

    # Nếu đổi sang tab khác hoặc chưa mở file nào, tiến hành mở file mới
    _workbook_path = get_type_scenario_path(scenario_type)

    # Load dữ liệu thuần túy (data_only=True) để lấy text, tránh lấy công thức Excel
    _workbook = openpyxl.load_workbook(_workbook_path, data_only=True)
    _current_loaded_type = scenario_type

    if config.SCENARIO_SHEET in _workbook.sheetnames:
        _sheet = _workbook[config.SCENARIO_SHEET]
    else:
        _sheet = _workbook.active

    return _sheet


# Load Data
# ==========================================================
def show_scenario(scenario_type):
    """
    Hàm chuẩn nhất để UI gọi sang. Vừa mở workbook đồng bộ biến, vừa trả về data.
    """
    sheet = open_workbook(scenario_type)
    return load_scenario_data(sheet)


def load_scenario_data(sheet):
    """
    Đọc dữ liệu an toàn từ sheet, bọc kiểm tra index chống lỗi file Excel thiếu cột.
    """
    scenario_list = []
    if sheet is None:
        return scenario_list

    for row in sheet.iter_rows(min_row=2, values_only=True):
        # Bỏ qua dòng trống hoàn toàn
        if not row or not any(cell is not None for cell in row):
            continue

        scene = {
            "scene_id": row[0] if len(row) > 0 and row[0] is not None else "",
            "template_key": row[1] if len(row) > 1 and row[1] is not None else "",
            "prompt_core": row[2] if len(row) > 2 and row[2] is not None else "",
            "character_list": row[3] if len(row) > 3 and row[3] is not None else "",
            "reference_list": row[4] if len(row) > 4 and row[4] is not None else "",
            "final_prompt": row[5] if len(row) > 5 and row[5] is not None else "",
            "status_time": row[6] if len(row) > 6 and row[6] is not None else "",
            "output_id": row[7] if len(row) > 7 and row[7] is not None else "",
            "file_name": row[8] if len(row) > 8 and row[8] is not None else ""
        }
        scenario_list.append(scene)

    return scenario_list


# Save Data
# ==========================================================
import openpyxl
import config


# Save Data
# ==========================================================
import openpyxl
import config


def save_specific_columns_from_ui(scenario_type, ui_data):
    """
    Mở file Excel gốc, chỉ cập nhật 3 cột: final_prompt, status_time, file_name
    dựa trên dữ liệu truyền xuống từ UI. Giữ nguyên tất cả các cột và các tab khác.
    """
    # 1. Xác định đúng đường dẫn file
    if scenario_type == config.SCENARIO_IMAGE:
        file_path = config.IMAGE_SCENARIO_FILE
    elif scenario_type == config.SCENARIO_VIDEO:
        file_path = config.VIDEO_SCENARIO_FILE
    else:
        raise ValueError(f"Loại scenario không hợp lệ: {scenario_type}")

    if not file_path.exists():
        raise FileNotFoundError(f"Không tìm thấy file Excel tại: {file_path}")

    try:
        # 2. Mở workbook ở chế độ GHI (KHÔNG dùng data_only=True để bảo toàn file)
        wb = openpyxl.load_workbook(file_path)

        if config.SCENARIO_SHEET in wb.sheetnames:
            ws = wb[config.SCENARIO_SHEET]
        else:
            ws = wb.active

        # Lấy số cột từ cấu hình EXCEL_COLUMNS của bạn
        col_final_prompt = config.EXCEL_COLUMNS["final_prompt"]
        col_status_time = config.EXCEL_COLUMNS["status_time"]
        col_file_name = config.EXCEL_COLUMNS["file_name"]

        # 3. Duyệt qua từng dòng dữ liệu từ UI gửi xuống
        # Dòng dữ liệu trong Excel bắt đầu từ dòng 2 (dòng 1 là tiêu đề)
        for index, row_values in enumerate(ui_data):
            current_row = index + 2

            # Đảm bảo dòng dữ liệu từ UI có đủ các phần tử để bốc tách
            # Thứ tự phần tử trong row_values tương ứng với thứ tự config.TREE_COLUMNS của bạn:
            # 0: scene_id, 1: template_key, 2: prompt_core, 3: character_list,
            # 4: reference_list, 5: final_prompt, 6: status_time, 7: output_id, 8: file_name

            if len(row_values) > 8:
                # Ghi đè duy nhất 3 cột mục tiêu
                ws.cell(row=current_row, column=col_final_prompt).value = row_values[5]
                ws.cell(row=current_row, column=col_status_time).value = row_values[6]
                ws.cell(row=current_row, column=col_file_name).value = row_values[8]

        # 4. Lưu lại file và giải phóng bộ nhớ
        wb.save(file_path)
        wb.close()

        # 5. Đồng bộ lại các biến tạm toàn cục dùng để đọc dữ liệu nếu cần
        global _workbook, _sheet
        _workbook = openpyxl.load_workbook(file_path, data_only=True)
        _sheet = _workbook[config.SCENARIO_SHEET]

        print(f"Cập nhật thành công 3 cột cho file: {file_path.name}")
        return True

    except PermissionError:
        raise RuntimeError(
            f"Không thể ghi! File {file_path.name} vẫn đang bị mở hoặc khóa bởi ứng dụng khác.\n"
            f"Vui lòng tắt hoàn toàn file Excel đó đi rồi thử lại."
        )
    except Exception as e:
        raise RuntimeError(f"Lỗi khi cập nhật dữ liệu Excel: {str(e)}")


# Update Cells
# ==========================================================
def update_status(row_index, status):
    if not _sheet:
        return
    _sheet.cell(row=row_index, column=config.EXCEL_COLUMNS["status_time"]).value = status


def update_filename(row_index, file_name):
    if not _sheet:
        return
    _sheet.cell(row=row_index, column=config.EXCEL_COLUMNS["file_name"]).value = file_name


def update_output_id(row_index, output_id):
    if not _sheet:
        return
    _sheet.cell(row=row_index, column=config.EXCEL_COLUMNS["output_id"]).value = output_id


def update_row(row_index, status=None, output_id=None, file_name=None):
    if status is not None:
        update_status(row_index, status)
    if output_id is not None:
        update_output_id(row_index, output_id)
    if file_name is not None:
        update_filename(row_index, file_name)


# Helpers
# ==========================================================
def get_sheet():
    return _sheet


def get_row_count():
    if not _sheet:
        return 0
    return _sheet.max_row - 1


def get_template_dict():
    """
    Đọc bảng tra cứu mẫu. Đảm bảo vượt qua kiểm tra _workbook an toàn.
    """
    global _workbook
    if not _workbook:
        return {}

    if "Templates" not in _workbook.sheetnames:
        return {}

    sheet = _workbook["Templates"]
    template_dict = {}
    for row in range(2, sheet.max_row + 1):
        key = sheet.cell(row=row, column=1).value
        if key:
            template_dict[str(key).strip().lower()] = str(
                sheet.cell(row=row, column=2).value or ""
            )
    return template_dict
